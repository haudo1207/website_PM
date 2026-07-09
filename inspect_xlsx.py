import openpyxl

wb = openpyxl.load_workbook("Skill_Master_From_Sheet.xlsx")
sheet = wb.active
print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

# Print first 3 rows of all columns
for r in [1, 2, 3]:
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    print(f"Row {r}: {row_vals}")
