import openpyxl
import json

wb = openpyxl.load_workbook("Skill_Master_From_Sheet.xlsx")
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

categories = {}
# Row 1 is Categories
# Row 3 is Groups
# Row 4+ is Skills

# Let's map column index (0-indexed) to category and group
col_mapping = {}

for col_idx in range(1, max_col + 1):
    cat_val = sheet.cell(1, col_idx).value
    group_val = sheet.cell(3, col_idx).value
    
    if cat_val is not None:
        cat_val = cat_val.strip()
    if group_val is not None:
        group_val = group_val.strip()
        
    if cat_val or group_val:
        col_mapping[col_idx] = {
            "category": cat_val,
            "group": group_val
        }

# Fill missing category names if they span columns
last_cat = None
for col_idx in sorted(col_mapping.keys()):
    if col_mapping[col_idx]["category"]:
        last_cat = col_mapping[col_idx]["category"]
    else:
        col_mapping[col_idx]["category"] = last_cat

# Print the columns config
print("Columns mapping:")
for col, mapping in col_mapping.items():
    print(f"Col {col}: Category='{mapping['category']}', Group='{mapping['group']}'")

# Now extract skills
skills_by_group = {} # (category, group) -> list of skills
for r in range(4, max_row + 1):
    has_data = False
    for col_idx in col_mapping.keys():
        val = sheet.cell(r, col_idx).value
        if val is not None:
            val = str(val).strip()
            if val:
                has_data = True
                mapping = col_mapping[col_idx]
                key = (mapping["category"], mapping["group"])
                if key not in skills_by_group:
                    skills_by_group[key] = []
                if val not in skills_by_group[key]:
                    skills_by_group[key].append(val)
    if not has_data:
        # Stop reading if we hit multiple empty rows
        pass

print("\nSkills Count by Group:")
for key, skills in skills_by_group.items():
    print(f"Category: {key[0]} | Group: {key[1]} | Skills: {len(skills)}")
