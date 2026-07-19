from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from .routers import auth, users, settings_router, skills, system_categories, members, meetings as meetings_router
from .routers import projects as projects_router
from .routers import task_groups as task_groups_router
from .routers import leave_requests as leave_requests_router
from .database import engine, Base
from .models import user, setting, member, skill_master, system_category, meeting as meeting_model
from .models import project as project_model
from .models import phase as phase_model
from .models import task_group as task_group_model
from .models import task as task_model
from .models import leave_request as leave_request_model
from sqlalchemy import text

# Create all tables (new v5 schema)
Base.metadata.create_all(bind=engine)

def apply_schema_updates():
    """Small idempotent migrations for existing installs without Alembic."""
    with engine.begin() as conn:
        conn.execute(text(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS data_scope VARCHAR(50) "
            "NOT NULL DEFAULT 'infrastructure'"
        ))
        conn.execute(text(
            "ALTER TABLE projects ADD COLUMN IF NOT EXISTS data_scope VARCHAR(50) "
            "NOT NULL DEFAULT 'infrastructure'"
        ))
        conn.execute(text(
            "CREATE INDEX IF NOT EXISTS ix_projects_data_scope ON projects (data_scope)"
        ))
        conn.execute(text(
            "UPDATE users SET data_scope = 'all' WHERE role = 'admin' AND data_scope = 'infrastructure'"
        ))
        # The legacy seed inserted users.id=1 explicitly and left the sequence at 1.
        conn.execute(text(
            "SELECT setval(pg_get_serial_sequence('users', 'id'), "
            "GREATEST(COALESCE((SELECT MAX(id) FROM users), 1), 1), true)"
        ))

apply_schema_updates()

def init_db_defaults():
    from .database import SessionLocal
    from .models.user import User
    from .models.setting import Setting
    from .utils.auth import hash_password
    import json
    import os

    db = SessionLocal()
    try:
        # 1. Initialize settings if missing
        if not db.query(Setting).filter(Setting.key == "column_config").first():
            cols = {
                "cols": ["DETAIL TASK", "PRIORITY", "MANDAY (EST)", "STATUS", "ASSIGNED"],
                "tab_names": []
            }
            db.add(Setting(key="column_config", value=json.dumps(cols, ensure_ascii=False)))

        if not db.query(Setting).filter(Setting.key == "policy").first():
            policy = {
                "rules": [
                    {
                        "field": "PRIORITY",
                        "value": "URGENT",
                        "manday_max": 2.0,
                        "min_words": 10,
                        "required_fields": ["ASSIGNED", "STATUS"]
                    },
                    {
                        "field": "PRIORITY",
                        "value": "CRITICAL",
                        "manday_max": 3.0,
                        "min_words": 15,
                        "required_fields": ["ASSIGNED", "STATUS"]
                    },
                    {
                        "field": "PRIORITY",
                        "value": "HIGH",
                        "manday_max": 5.0,
                        "min_words": 5,
                        "required_fields": ["ASSIGNED"]
                    }
                ]
            }
            db.add(Setting(key="policy", value=json.dumps(policy, ensure_ascii=False)))

        if not db.query(Setting).filter(Setting.key == "ai_config").first():
            ai = {
                "base_url": os.environ.get("AI_BASE_URL", "https://api.shopaikey.com/v1"),
                "api_key": os.environ.get("AI_API_KEY", ""),
                "model": os.environ.get("AI_MODEL", "gpt-4o-mini"),
                "system_prompt": (
                    "Bạn là trợ lý ảo kiểm soát chất lượng & tuân thủ quy trình dự án.\n"
                    "Hãy đánh giá dòng công việc (task) sau từ Google Sheet:\n"
                    "1. Kiểm tra xem mô tả công việc (DETAIL TASK) có đủ rõ ràng, có đầy đủ mục tiêu, kết quả bàn giao hay không.\n"
                    "2. Đánh giá xem Manday (MANDAY (EST)) có quá cao hay quá thấp so với mô tả công việc không.\n"
                    "3. Trả về phán quyết (PASS, FAIL, hoặc REVIEW).\n"
                    "4. Đưa ra Lý do (tiếng Việt).\n"
                    "5. Đưa ra Gợi ý cải thiện cụ thể (tiếng Việt).\n"
                    "Định dạng trả về bắt buộc là JSON hợp lệ có dạng: {\"verdict\":\"...\",\"reason\":\"...\",\"suggestion\":\"...\"}\n"
                    "Không viết markdown, không thêm ký tự ngoài JSON."
                ),
                "check_interval_hours": 1
            }
            db.add(Setting(key="ai_config", value=json.dumps(ai, ensure_ascii=False)))

        # 2. Initialize default admin user if no users exist
        if db.query(User).count() == 0:
            admin = User(
                id=1,
                email="admin@company.com",
                full_name="Admin Company",
                hashed_pw=hash_password("admin123"),
                role="admin",
                is_active=True
            )
            db.add(admin)
            print("[*] Created default admin: admin@company.com")

        # 3. Seed Skills from Excel if empty
        from .models.skill_master import Category, Group, Skill
        if db.query(Category).count() == 0:
            import openpyxl
            xlsx_path = "Skill_Master_From_Sheet.xlsx"
            if os.path.exists(xlsx_path):
                wb = openpyxl.load_workbook(xlsx_path)
                sheet = wb.active
                max_row = sheet.max_row
                max_col = sheet.max_column

                col_mapping = {}
                for col_idx in range(1, max_col + 1):
                    cat_val = sheet.cell(1, col_idx).value
                    group_val = sheet.cell(3, col_idx).value
                    if cat_val is not None:
                        cat_val = str(cat_val).strip()
                    if group_val is not None:
                        group_val = str(group_val).strip()
                    if cat_val or group_val:
                        col_mapping[col_idx] = {"category": cat_val, "group": group_val}

                last_cat = None
                for col_idx in sorted(col_mapping.keys()):
                    if col_mapping[col_idx]["category"]:
                        last_cat = col_mapping[col_idx]["category"]
                    else:
                        col_mapping[col_idx]["category"] = last_cat

                db_categories = {}
                db_groups = {}

                for col_idx, mapping in col_mapping.items():
                    cat_name = mapping["category"]
                    grp_name = mapping["group"]
                    if not cat_name or not grp_name:
                        continue
                    if cat_name not in db_categories:
                        cat_obj = Category(name=cat_name, is_active=True)
                        db.add(cat_obj)
                        db.flush()
                        db_categories[cat_name] = cat_obj
                    grp_key = (cat_name, grp_name)
                    if grp_key not in db_groups:
                        grp_obj = Group(name=grp_name, category_id=db_categories[cat_name].id)
                        db.add(grp_obj)
                        db.flush()
                        db_groups[grp_key] = grp_obj

                for r in range(4, max_row + 1):
                    for col_idx in col_mapping.keys():
                        val = sheet.cell(r, col_idx).value
                        if val is not None:
                            val = str(val).strip()
                            if val:
                                mapping = col_mapping[col_idx]
                                cat_name = mapping["category"]
                                grp_name = mapping["group"]
                                if not cat_name or not grp_name:
                                    continue
                                grp_obj = db_groups.get((cat_name, grp_name))
                                if grp_obj:
                                    existing = db.query(Skill).filter(Skill.name == val, Skill.group_id == grp_obj.id).first()
                                    if not existing:
                                        db.add(Skill(name=val, group_id=grp_obj.id, is_active=True))
                print("[*] Seeded skills from excel")

        # 4. Seed system categories
        from .models.system_category import Position, Department, TaskPriority, TaskStatus, Team, Customer
        if db.query(Position).count() == 0:
            for name in ["PM", "Leader", "Developer", "Tester", "Intern"]:
                db.add(Position(name=name))

        if db.query(Department).count() == 0:
            for name in ["AI", "ERP", "QA", "Infrastructure", "Marketing"]:
                db.add(Department(name=name))

        if db.query(Team).count() == 0:
            for name in ["Sales", "Presales", "Technical", "Intern L2", "Intern L1 Tech",
                          "Developer", "Marketing", "Freelancer", "Back Office", "Intern L1"]:
                db.add(Team(name=name))

        if db.query(TaskPriority).count() == 0:
            priorities = [
                {"name": "Normal", "kpi_base": 6, "color": "🟢 Xanh"},
                {"name": "High", "kpi_base": 12, "color": "🟠 Cam"},
                {"name": "Critical", "kpi_base": 20, "color": "🔴 Đỏ"},
                {"name": "Interrupt", "kpi_base": 6, "color": "🟣 Tím"}
            ]
            for p in priorities:
                db.add(TaskPriority(name=p["name"], kpi_base=p["kpi_base"], color=p["color"]))

        if db.query(TaskStatus).count() == 0:
            statuses = ["Waiting", "Process", "Done", "Cancel", "Rework", "To Do"]
            for name in statuses:
                db.add(TaskStatus(name=name))

        if db.query(Customer).count() == 0:
            customers = ["Samsung SDS", "LG CNS", "Viettel", "FPT Software", "Vingroup"]
            for name in customers:
                db.add(Customer(name=name))

        # 5. Seed default platforms
        from .models.project import Platform
        if db.query(Platform).count() == 0:
            default_platforms = [
                {"name": "Telegram", "icon": "📨", "color": "#0088cc"},
                {"name": "Zalo", "icon": "💬", "color": "#0068ff"},
                {"name": "Slack", "icon": "💼", "color": "#4A154B"},
                {"name": "Microsoft Teams", "icon": "🟦", "color": "#6264A7"},
                {"name": "Discord", "icon": "🎮", "color": "#5865F2"},
                {"name": "WhatsApp", "icon": "📱", "color": "#25D366"},
                {"name": "Khác", "icon": "🔗", "color": "#6b7280"},
            ]
            for p in default_platforms:
                db.add(Platform(name=p["name"], icon=p["icon"], color=p["color"], is_active=True))
            print("[*] Seeded default platforms")

        # 6. Seed initial members
        from .models.member import Member
        if db.query(Member).count() == 0:
            initial_members = [
                {"display_name": "1.Minhpn",  "full_name": "Minh",   "telegram_username": "@minhuit911",   "team": "Sales"},
                {"display_name": "6.Phil",     "full_name": "Phil",   "telegram_username": "@phile0920",    "team": "Sales"},
                {"display_name": "8.Nhớlnha",  "full_name": "Nhớ",   "telegram_username": "@AnhNho66",     "team": "Sales"},
                {"display_name": "9.Phátnt",   "full_name": "Phát",  "telegram_username": "@phatnguyeen",  "team": "Sales"},
                {"display_name": "2.Phướcpv",  "full_name": "Phước", "telegram_username": "@nessivu",      "team": "Presales"},
                {"display_name": "3.Hând",     "full_name": "Hân",   "telegram_username": "@handiep",      "team": "Presales"},
                {"display_name": "4.Yêmh",     "full_name": "Yên",   "telegram_username": "@yemhoang96",   "team": "Presales"},
                {"display_name": "5.Khôipm",   "full_name": "Khôi",  "telegram_username": "@BlackDrag0n98","team": "Presales"},
                {"display_name": "7.Vypt",     "full_name": "Vy",    "telegram_username": "@vyvyzz13",     "team": "Technical"},
                {"display_name": "10.Huyht",   "full_name": "Huy",   "telegram_username": "@ht_hyu",       "team": "Technical"},
                {"display_name": "11.Hânlnl",  "full_name": "Hân",   "telegram_username": "@Linhhan0399",  "team": "Technical"},
                {"display_name": "12.Vũnl",    "full_name": "Vũ",    "telegram_username": "@vu_nguyen27",  "team": "Technical"},
                {"display_name": "14.Túbda",   "full_name": "Tú",    "telegram_username": "@numm0204",     "team": "Technical"},
                {"display_name": "15.Huấnnv",  "full_name": "Huấn",  "telegram_username": "@huannv21",     "team": "Technical"},
                {"display_name": "17.Sơnpn",   "full_name": "Sơn",   "telegram_username": "@Sonngocpham",  "team": "Technical"},
                {"display_name": "18.Thuậnlh", "full_name": "Thuận", "telegram_username": "@thuanle0608",  "team": "Technical"},
                {"display_name": "25.Chínhvđ", "full_name": "Chính", "telegram_username": "@duci59",       "team": "Technical"},
                {"display_name": "28.Phátntu", "full_name": "Phát",  "telegram_username": "@phatnguyentuan","team": "Technical"},
                {"display_name": "13.Tâylt",   "full_name": "Tây",   "telegram_username": "@lethitay",     "team": "Intern L2"},
                {"display_name": "16.Dũngnh",  "full_name": "Dũng",  "telegram_username": "@ru4ff",        "team": "Intern L2"},
                {"display_name": "26.Phúlt",   "full_name": "Phú",   "telegram_username": "@letrieuphu",   "team": "Intern L2"},
                {"display_name": "27.Duynđ",   "full_name": "Duy",   "telegram_username": "@duyisme_rynn", "team": "Intern L2"},
                {"display_name": "29.Anhpxt",  "full_name": "Anh",   "telegram_username": "@tunanhdapoet",  "team": "Intern L2"},
                {"display_name": "30.Tàipv",   "full_name": "Tài",   "telegram_username": "@taipvan",      "team": "Intern L2"},
                {"display_name": "31.Hiệplh",  "full_name": "Hiệp",  "telegram_username": "@hhieple03",    "team": "Intern L2"},
                {"display_name": "32.Thànhndn","full_name": "Thành", "telegram_username": "@thanh25324",   "team": "Intern L2"},
                {"display_name": "33.Phươngvt","full_name": "Phương","telegram_username": "@tpunn00",      "team": "Intern L2"},
                {"display_name": "Sếp Thương", "full_name": "Thương","telegram_username": None,  "team": "Developer"},
                {"display_name": "Sếp Nhi",    "full_name": "Nhi",   "telegram_username": None,  "team": "Developer"},
                {"display_name": "Chị Trang",  "full_name": "Trang", "telegram_username": "@Trang07",       "team": "Developer"},
            ]
            for data in initial_members:
                try:
                    mb = Member(
                        display_name=data["display_name"],
                        full_name=data["full_name"],
                        telegram_username=data.get("telegram_username"),
                        team=data.get("team")
                    )
                    db.add(mb)
                    db.flush()
                except Exception:
                    db.rollback()
            print("[*] Seeded initial members")

        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[!] Error seeding database: {e}")
    finally:
        db.close()

# Run database auto-seeding
init_db_defaults()


app = FastAPI(
    title="Task Compliance Portal",
    version="5.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

app.include_router(auth.router,                prefix="/api/auth")
app.include_router(users.router,               prefix="/api/users")
app.include_router(projects_router.router,     prefix="/api")
app.include_router(task_groups_router.router,   prefix="/api")
app.include_router(settings_router.router,     prefix="/api/settings")
app.include_router(skills.router,              prefix="/api/skills")
app.include_router(system_categories.router,   prefix="/api/system-categories")
app.include_router(members.router,             prefix="/api/members")
app.include_router(leave_requests_router.router, prefix="/api/leave-requests")
app.include_router(meetings_router.router,    prefix="/api/meetings")

@app.get("/health")
@app.get("/api/health")
def health():
    return {"status": "ok", "version": "5.0.0"}
